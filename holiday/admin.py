import datetime
from tempfile import NamedTemporaryFile

from django.contrib import admin, messages
from django.db.models import Prefetch
from django.http import HttpResponse
from django_better_admin_arrayfield.admin.mixins import DynamicArrayMixin
from django.utils.translation import gettext_lazy as _
from django_better_admin_arrayfield.forms.fields import DynamicArrayField
from django_better_admin_arrayfield.forms.widgets import DynamicArrayWidget
from openpyxl import Workbook
from openpyxl.styles import Alignment
from django import forms
from ordered_model.admin import (
    OrderedStackedInline,
    OrderedInlineModelAdminMixin,
    OrderedModelAdmin,
)

from holiday.models import (
    Holiday,
    HolidaySection,
    Registration,
    Outing,
    SectionProgram,
    _send_registration_notification,
)
from members.models import User


class DynamicArrayDateInputWidget(DynamicArrayWidget):
    def __init__(self, *args, **kwargs):
        kwargs["subwidget_form"] = forms.DateInput
        super().__init__(*args, **kwargs)


class OutingInline(admin.StackedInline):
    model = Outing
    extra = 0


class SectionProgramForm(forms.ModelForm):
    animateur = forms.ModelMultipleChoiceField(
        queryset=User.objects.none(),  # will set both qs & choices in __init__
        required=False,
    )

    class Meta:
        model = SectionProgram
        fields = "__all__"

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)

        # on first init (per process), pull & cache all staff
        cls = self.__class__
        if not hasattr(cls, "_staff_cache"):
            cls._staff_cache = list(User.objects.filter(is_staff=True))

        staff_list = cls._staff_cache
        staff_ids = [u.pk for u in staff_list]

        # set queryset (for validation) and choices (for rendering)
        qs = User.objects.filter(pk__in=staff_ids)
        self.fields["animateur"].queryset = qs
        # override the choices so Django doesn’t re-evaluate the qs on .choices
        self.fields["animateur"].choices = [(u.pk, str(u)) for u in staff_list]


class SectionProgramInline(admin.StackedInline):
    # def get_queryset(self, request):
    #     qs = super().get_queryset(request)
    #     # this will pull all animateurs for *all* inline rows in one query
    #     return qs.prefetch_related("animateur")

    def get_formset(self, request, obj=None, **kwargs):
        # 1) build your formset class
        FormSet = super().get_formset(request, obj, **kwargs)

        # 2) load all staff ONCE and cache it on the class
        if not hasattr(self, "_staff_cache"):
            staff = list(User.objects.filter(is_staff=True))  # ← 1 DB hit
            pks = [u.pk for u in staff]
            choices = [(u.pk, str(u)) for u in staff]
            self._staff_cache = (pks, choices)

        pks, choices = self._staff_cache

        # 3) override the base_fields of the form
        bf = FormSet.form.base_fields["animateur"]
        # validation will still work off a pk__in lookup, but we never __evaluate__ again
        bf.queryset = User.objects.filter(pk__in=pks)
        bf.choices = choices

        return FormSet

    # autocomplete_fields = ("animateur",)
    # form = SectionProgramForm
    def get_queryset(self, request):
        qs = super().get_queryset(request)
        # Pull every animateur for all inline rows in one shot:
        return qs.prefetch_related("animateur")

    model = SectionProgram
    extra = 0


@admin.register(HolidaySection)
class HolidaySectionAdmin(admin.ModelAdmin):
    model = HolidaySection
    extra = 0
    inlines = [SectionProgramInline, OutingInline]
    list_display = ["section", "holiday", "capacity", "remaining_capacity"]
    list_filter = ["section", "holiday", "capacity"]

    readonly_fields = ["remaining_capacity_table"]

    fieldsets = [
        [None, {"fields": ["section", "holiday", "capacity", "description"]}],
        [_("remaining capacity"), {"fields": ["remaining_capacity_table"]}],
    ]

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        return (
            qs
            # so .section and .holiday don’t hit the DB
            .select_related("section", "holiday")
            # pull in every Registration for each holiday…
            .prefetch_related(
                Prefetch(
                    "holiday__registration_set",
                    queryset=Registration.objects.all(),
                    to_attr="all_regs_for_holiday",
                )
            )
        )


# Register your models here.
@admin.register(Holiday)
class HolidayAdmin(admin.ModelAdmin, DynamicArrayMixin):
    model = Holiday
    list_display = ["name", "start_date", "end_date", "price", "registration_open"]

    actions = ["export_registration"]

    def formfield_for_dbfield(self, db_field, request, **kwargs):
        if db_field.name == "blacklisted_dates":
            kwargs["widget"] = DynamicArrayDateInputWidget()
        return super().formfield_for_dbfield(db_field, request, **kwargs)

    formfield_overrides = {DynamicArrayField: {"widget": DynamicArrayDateInputWidget}}

    @admin.action(description=_("export_registration_desc"))
    def export_registration(self, request, queryset):
        if queryset.count() > 1:
            self.message_user(
                request,
                "L'export ne fonctionne que pour une vacances à la fois",
                messages.ERROR,
            )
            return
        holiday = queryset.first()
        wb = Workbook()

        ws = wb.active
        ws.title = "Inscriptions"
        # number of days
        start_date = holiday.start_date
        end_date = holiday.end_date

        ws.cell(1, 1, value=f"{holiday} ({start_date} - {end_date})").alignment = (
            Alignment(horizontal="center", vertical="center")
        )
        ws.cell(2, 1, value="Enfant").alignment = Alignment(
            horizontal="center", vertical="center"
        )
        ws.cell(3, 1, value="Prénom").alignment = Alignment(
            horizontal="center", vertical="center"
        )
        ws.cell(3, 2, value="Nom").alignment = Alignment(
            horizontal="center", vertical="center"
        )
        ws.cell(3, 3, value="Date de Naissance").alignment = Alignment(
            horizontal="center", vertical="center"
        )
        ws.cell(3, 4, value="Allergies").alignment = Alignment(
            horizontal="center", vertical="center"
        )
        ws.merge_cells(start_row=2, end_row=2, start_column=1, end_column=4)
        ws.cell(2, 5, value="Parent").alignment = Alignment(
            horizontal="center", vertical="center"
        )
        ws.cell(3, 5, value="Prénom").alignment = Alignment(
            horizontal="center", vertical="center"
        )
        ws.cell(3, 6, value="Nom").alignment = Alignment(
            horizontal="center", vertical="center"
        )
        ws.cell(3, 7, value="Email").alignment = Alignment(
            horizontal="center", vertical="center"
        )
        ws.merge_cells(start_row=2, end_row=2, start_column=5, end_column=7)
        ws.cell(2, 8, value="Section").alignment = Alignment(
            horizontal="center", vertical="center"
        )
        ws.merge_cells(start_row=2, end_row=3, start_column=8, end_column=8)
        ws.cell(2, 9, value="# jours").alignment = Alignment(
            horizontal="center", vertical="center"
        )
        ws.merge_cells(start_row=2, end_row=3, start_column=9, end_column=9)
        prev_date = holiday.start_date
        num_col = 10
        while prev_date <= holiday.end_date:
            if prev_date.weekday() > 4:
                # weekend
                prev_date = prev_date + datetime.timedelta(days=1)
                continue
            ws.cell(2, num_col, value=prev_date.strftime("%A"))
            ws.cell(3, num_col, value=prev_date.strftime("%d-%m-%Y"))
            prev_date = prev_date + datetime.timedelta(days=1)
            num_col += 1
        ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=num_col)
        num_row = 4
        for registration in holiday.registration_set.all():
            num_col = 1
            # Child Info
            ws.cell(num_row, num_col, value=registration.child.first_name)
            num_col += 1
            ws.cell(num_row, num_col, value=registration.child.last_name)
            num_col += 1
            ws.cell(
                num_row,
                num_col,
                value=registration.child.birth_date.strftime("%d-%m-%Y"),
            )
            num_col += 1
            ws.cell(num_row, num_col, value=registration.notes)
            num_col += 1
            ws.cell(num_row, num_col, value=registration.child.parent.first_name)
            num_col += 1
            ws.cell(num_row, num_col, value=registration.child.parent.last_name)
            num_col += 1
            ws.cell(num_row, num_col, value=registration.child.parent.email)
            num_col += 1
            ws.cell(num_row, num_col, value=registration.section.name)
            num_col += 1
            ws.cell(num_row, num_col, value=registration.number_of_days)
            num_col += 1
            prev_date = holiday.start_date
            dates = list(registration.dates)
            while prev_date <= holiday.end_date:
                if prev_date.weekday() > 4:
                    # weekend
                    prev_date = prev_date + datetime.timedelta(days=1)
                    continue
                if prev_date in dates:
                    ws.cell(num_row, num_col, value=1)
                prev_date = prev_date + datetime.timedelta(days=1)
                num_col += 1
            num_row += 1
        with NamedTemporaryFile() as tmp:
            wb.save(tmp.name)
            tmp.seek(0)
            response = HttpResponse(
                tmp,
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response["Content-Disposition"] = "attachment; filename=inscriptions.xlsx"
            return response


@admin.register(Registration)
class RegistrationAdmin(admin.ModelAdmin, DynamicArrayMixin):
    model = Registration
    list_display = ["holiday", "_child", "number_of_days", "section", "status", "cost"]

    @admin.display(description=_("child"))
    def _child(self, obj):
        return f"{obj.child.first_name} {obj.child.last_name}"

    list_filter = ["holiday", "child__parent", "status", "section"]
    actions = ["resend_email"]

    list_per_page = 20

    @admin.action(description=_("resend_email"))
    def resend_email(self, request, queryset):
        for registration in queryset.all():
            _send_registration_notification(registration)
